import pyodbc
import threading
import xlwt
import time
import smtplib
from email.header import Header
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import threadpool
import openpyxl


class SendEmail:
    def __init__(self, sender, receivers, mail_title, path, filename, smtpserver, username, password):
        self.sender = sender
        self.receivers = receivers
        self.mail_title = mail_title
        self.path = path
        self.filename = filename
        self.smtpserver = smtpserver
        self.username = username
        self.password = password

    def createMessage(self):
        # 创建一个带附件的实例
        message = MIMEMultipart()
        message['From'] = self.sender
        message['To'] = self.receivers
        message['Subject'] = Header(self.mail_title, 'utf-8')
        # 邮件正文内容
        message.attach(MIMEText('请查收附件数据 from 弘康！', 'plain', 'utf-8'))
        part = MIMEBase('application', 'octet-stream')
        # 此处的path是需要添加附件的文件路径
        part.set_payload(open(self.path, "rb").read())
        encoders.encode_base64(part)
        # filename表示附件中文件的命名
        part.add_header('Content-Disposition', 'attachment', filename=self.filename)
        message.attach(part)
        return message

    def senEmail(self, message: MIMEMultipart):
        smtpObj = smtplib.SMTP_SSL(self.smtpserver)
        smtpObj.connect(self.smtpserver)
        smtpObj.login(self.username, self.password)
        smtpObj.sendmail(self.sender, self.receivers.split(";"), message.as_string())


class ProcessThread(threading.Thread):
    # threadID指定线程id
    # name指定线程名称
    # config_row指定线程处理的配置表中的一行
    # config_desc指定线程处理的配置表的列名
    # sender指定发送邮件的用户
    # receivers指定接收邮件的用户（以列表形式传递）
    # mail_title指定发送邮件的邮件头
    # path指定数据文件的路径
    def __init__(self,threadID, name, config_row,
                 sender, path, smtpserver, username, password, run_flag):
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.config_row = config_row
        self.sender = sender
        self.path = path
        self.smtpserver = smtpserver
        self.username = username
        self.password = password
        self.run_flag = run_flag

    # 写出的数据文件名为{path}{title}_{ticks} 文件的保存路径,path必须以/结尾
    # row是配置表中的一行数据
    # columns_desc是配置表中的列名列表
    def run(self):
        # 从config_row中读取出dsn\sql\emails\title
        taskId = str(self.config_row[0]).strip()
        dsn = str(self.config_row[1]).strip()
        sql = str(self.config_row[2]).strip()
        receivers = str(self.config_row[3]).strip()
        title = str(self.config_row[4]).strip()
        # 如果dsn\sql\emails不为空,则继续执行
        if dsn is not None and sql is not None and receivers is not None and title is not None:
            # 建立连接并读取数据
            conn = pyodbc.connect(dsn=dsn)
            cur = conn.cursor()
            cur.execute(sql)
            result = cur.fetchall()
            # 创建一个excel
            book = openpyxl.Workbook()
            sheet = book.create_sheet(index=0)
            # 将读取到的数据写入创建的excel中
            # 将字段名称写入excel
            for col, field in enumerate([field[0] for field in cur.description]):
                sheet.cell(1, col + 1, field)
            row_index = 2
            # 将每一行的数据遍历写入excel中
            for row in result:
                for col_index, field_value in enumerate(row):
                    sheet.cell(row_index, col_index + 1, field_value)
                row_index += 1

            # 将excel进行保存操作
            ticks = time.strftime("%Y%m%d%H%M%S", time.localtime())
            data_path = "{path}/{title}_{ticks}.xlsx".format(path=self.path, title=title, ticks=ticks)
            file_name = "{title}_{ticks}.xlsx".format(title=title, ticks=ticks)
            mail_title = "{title}_{ticks}".format(title=title, ticks=ticks)
            book.save(data_path)
            cur.close()
            conn.close()

            # 发送携带数据的邮件
            senEmailEntity = SendEmail(sender=self.sender, receivers=receivers,
                                       mail_title=mail_title, path=data_path,
                                       filename=file_name, smtpserver=self.smtpserver,
                                       username=self.username, password=self.password)
            message = senEmailEntity.createMessage()
            senEmailEntity.senEmail(message)
            print('发送数据成功:',taskId)
            conn = pyodbc.connect(dsn="spider")
            cur = conn.cursor()

            # 本次数据处理完之后修改数据的状态
            cur.execute("""update com_send_email_config a
                                            set a.if_only_once = 0
                                            where a.if_only_once = 1
                                            and a.is_running =  {run_flag}
                                            and a.id = {taskId}
                                            """.format(run_flag=self.run_flag,taskId=taskId))
            cur.commit()
            cur.execute("""update com_send_email_config a
                                        set a.is_running = 0
                                        where a.is_running = {run_flag}
                                        and a.id = {taskId}
                                        """.format(run_flag=self.run_flag,taskId=taskId))
            cur.commit()
            cur.close()
            conn.close()

def readConfig(tableName, is_running):
    conn = pyodbc.connect(dsn='spider')
    cur = conn.cursor()
    cur.execute("select id,dsn, sql_, emails receivers, title from {tableName} where is_running = {is_running} and is_del = 0".format(tableName=tableName,is_running=is_running))
    config_results = cur.fetchall()
    cur.close()
    conn.close()
    return config_results

if __name__ == '__main__':
    # 该连接对象用于读取com_send_email_config配置表的数据信息
    conn = pyodbc.connect(dsn="spider")
    cur = conn.cursor()
    while (True):
        # 标记本次执行的时间戳
        execute_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())

        # 获取一个数值(本次执行的唯一标识) 赋值给本次需要执行的sql
        cur.execute("""select nvl(max(a.is_running) + 1 ,1) max_running_value from com_send_email_config a""")
        run_flag = cur.fetchall()[0][0]

        # 以last_run_time为基准 根据hour\day\month的时间间隔,获取next_run_time的时间
        # 后续比较 current_timestamp 与 next_run_time
        # 刷新is_running = 0(表示没有在运行)数据的next_run_time字段
        cur.execute("""update com_send_email_config a
                        set a.next_run_time = date_add(date_add(date_add(date_add(date_add(a.last_run_time,interval nvl(a.second_,0) second),interval nvl(a.minute_,0) minute),interval nvl(a.hour_,0) hour),interval nvl(a.day_,0) day),interval nvl(a.month_,0) month)
                        where a.is_del = 0
                        and a.is_running = 0""")
        cur.execute("commit")

        # 刷新current_timestamp >= a.next_run_time数据的is_running和last_run_time字段(需要执行数据的sql 先提前修改状态)
        # 表示run_flag 表示本次需要执行的任务
        cur.execute("""update com_send_email_config a
                        set a.is_running = {run_flag},
                        a.last_run_time = a.next_run_time
                        where a.next_run_time <= str_to_date('{execute_time}','%Y-%m-%d %H:%i:%s')
                        and a.is_running = 0
                        and a.is_del = 0""".format(run_flag=run_flag, execute_time=execute_time))
        cur.execute("commit")

        # 额外将is_running状态为1的run_flag设置为本次需要执行的状态
        cur.execute("""update com_send_email_config a
                        set a.is_running = {run_flag}
                        where a.if_only_once = 1
                        and a.is_del = 0""".format(run_flag=run_flag))
        cur.execute("commit")

        # 使用一个单独的线程处理is_running = {run_flag} 的一条数据
        threads = list()
        config_results = readConfig(tableName="com_send_email_config", is_running=run_flag)

        # 创建线程对象并将线程对象添加到一个列表中，然后遍历列表执行线程
        id = 1
        for row in config_results:
            threads.append(ProcessThread(threadID=id, name="thread" + str(id) + str(run_flag),
                                        config_row=row, sender="from@hongkang-life.com",
                                        path="/opt/sendmail/data", smtpserver="smtp.qiye.163.com",
                                        username="from@hongkang-life.com",
                                        password="2017.com",
                                        run_flag=run_flag))
            id += 1
        print('threads:',len(threads))
        if len(threads) != 0:
            for thread in threads:
                thread.start()
        time.sleep(10)


